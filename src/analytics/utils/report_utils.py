import pandas as pd
import numpy as np
from datetime import datetime
from pathlib import Path
import logging
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

logger = logging.getLogger(__name__)

class ReportGenerator:    
    def __init__(self):
        self.styles = getSampleStyleSheet()
        self.title_style = self.styles['Heading1']
        self.heading_style = self.styles['Heading2']
        self.normal_style = self.styles['Normal']
        
        self.styles.add(ParagraphStyle(
            name='CenterTitle',
            parent=self.styles['Heading1'],
            alignment=1,  # Center alignment
            spaceAfter=30
        ))

    def generate_pdf_report(self, data_frames, titles, filename, 
                           orientation='portrait'):
        if hasattr(filename, '__fspath__'): 
            filename = str(filename)
        
        if orientation == 'landscape':
            page_size = landscape(A4)
        else:
            page_size = A4
        
        doc = SimpleDocTemplate(
            filename,
            pagesize=page_size,
            rightMargin=72,
            leftMargin=72,
            topMargin=72,
            bottomMargin=72
        )
        
        story = []
        
        title = Paragraph(f"CreditFlow360 - {datetime.now().strftime('%Y-%m-%d')}", 
                         self.styles['CenterTitle'])
        story.append(title)
        story.append(Spacer(1, 20))
        
        for df, title in zip(data_frames, titles):
            if df.empty:
                continue
                
            heading = Paragraph(title, self.styles['Heading2'])
            story.append(heading)
            story.append(Spacer(1, 12))
            
            data = [df.columns.tolist()] + df.values.tolist()
            
            table = Table(data, repeatRows=1)
            
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTSIZE', (0, 1), (-1, -1), 10),
                ('ALIGN', (1, 1), (-1, -1), 'RIGHT'),
            ]))
            
            story.append(table)
            story.append(Spacer(1, 20))
        
        doc.build(story)
        logger.info(f"✅ PDF report saved: {filename}")
    

    def generate_excel_report(self, data_frames, sheet_names, filename):
        if hasattr(filename, '__fspath__'): 
            filename = str(filename)
        
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            for df, sheet_name in zip(data_frames, sheet_names):
                if df.empty:
                    continue
                    
                df.to_excel(writer, sheet_name=sheet_name, index=False)
                
                workbook = writer.book
                worksheet = writer.sheets[sheet_name]
                
                header_font = Font(bold=True, color='FFFFFF')
                header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
                
                for cell in worksheet[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal='center')
                
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
                
                for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
                    for cell in row:
                        if isinstance(cell.value, (int, float)):
                            if cell.value > 1000:
                                cell.number_format = '₹#,##0'
                            elif cell.value < 1:
                                cell.number_format = '0.00%'
                            else:
                                cell.number_format = '#,##0'
        
        logger.info(f"✅ Excel report saved: {filename}")




    def generate_summary_stats(self, df, group_cols, metric_cols, agg_funcs=None):
        if agg_funcs is None:
            agg_funcs = ['count', 'sum', 'mean', 'min', 'max']
        
        summary = df.groupby(group_cols)[metric_cols].agg(agg_funcs).round(2)
        return summary